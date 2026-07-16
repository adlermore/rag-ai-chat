"""
Парсинг документов → список Chunk.

PDF/DOCX — через Docling (markdown-экспорт → структурный чанкинг).
XLSX — через openpyxl построчно (строка + шапка + имя листа), т.к. строка без
контекста бесполезна для поиска (docs/01-SPEC.md).

Docling — тяжёлая зависимость, импортируется лениво.
"""
from __future__ import annotations

from pathlib import Path

from .chunking import chunk_text, excel_row_chunk
from .models import Chunk

_SUPPORTED = {".pdf": "pdf", ".docx": "docx", ".xlsx": "xlsx"}


def _load_converter():
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions, TableFormerMode
    from docling.document_converter import DocumentConverter, PdfFormatOption

    # PDF заказчика цифровые (docs/01-SPEC.md) → OCR off (быстрее и точнее).
    pdf_opts = PdfPipelineOptions()
    pdf_opts.do_ocr = False
    # TableFormer в режиме FAST: ACCURATE (дефолт) на CPU катастрофически медленный
    # — многостраничный PDF мог обрабатываться десятки минут. FAST сохраняет
    # таблицы, но многократно быстрее (критично на self-hosted CPU-сервере).
    pdf_opts.do_table_structure = True
    pdf_opts.table_structure_options.mode = TableFormerMode.FAST
    return DocumentConverter(
        format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=pdf_opts)}
    )


def _parse_docling(path: Path, *, document_id: str, version: int, doc_type: str) -> list[Chunk]:
    converter = _load_converter()
    doc = converter.convert(str(path)).document
    markdown = doc.export_to_markdown()
    # Точный per-page split — открытый хвост (Docling markdown не несёт номеров
    # страниц напрямую); page=None, уточняется позже.
    return chunk_text(
        markdown,
        document_id=document_id,
        document_version=version,
        doc_type=doc_type,
        page=None,
    )


def _parse_xlsx(path: Path, *, document_id: str, version: int) -> list[Chunk]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    chunks: list[Chunk] = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue  # пустой лист
        headers = [str(h) if h is not None else "" for h in header]
        for r_idx, row in enumerate(rows, start=2):  # строка 1 — шапка
            values = [str(v) if v is not None else "" for v in row]
            if not any(v.strip() for v in values):
                continue  # пустая строка
            chunks.append(
                excel_row_chunk(
                    document_id=document_id,
                    document_version=version,
                    sheet=ws.title,
                    row=r_idx,
                    headers=headers,
                    values=values,
                )
            )
    wb.close()
    return chunks


def parse_document(
    path: str | Path, *, document_id: str, version: int
) -> list[Chunk]:
    """Парсит один файл в чанки с метаданными источника."""
    path = Path(path)
    doc_type = _SUPPORTED.get(path.suffix.lower())
    if not doc_type:
        raise ValueError(f"Неподдерживаемый формат: {path.suffix}")
    if doc_type == "xlsx":
        return _parse_xlsx(path, document_id=document_id, version=version)
    return _parse_docling(
        path, document_id=document_id, version=version, doc_type=doc_type
    )
